# scripts/process_tdms.py
import os
import pandas as pd
from nptdms import TdmsFile
from openpyxl import load_workbook
from datetime import datetime

def process_tdms_file(tdms_path, output_excel):
    freq_name = os.path.basename(tdms_path).replace(".tdms", "").strip()
    group_name = "accelerationgroup"

    try:
        tdms_file = TdmsFile.read(tdms_path)
        z_channel = tdms_file[group_name].channels()[0]
        z_data = z_channel[:]
    except Exception as e:
        print(f"[ERROR] Failed to read {tdms_path}: {e}")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    col_name = f"z-axis-{timestamp}"
    df_new = pd.DataFrame({col_name: z_data})

    if os.path.exists(output_excel):
        book = load_workbook(output_excel)
        writer = pd.ExcelWriter(output_excel, engine="openpyxl", mode='a', if_sheet_exists='overlay')
        writer.book = book

        if freq_name in book.sheetnames:
            existing_df = pd.read_excel(output_excel, sheet_name=freq_name)
            updated_df = pd.concat([existing_df, df_new], axis=1)
        else:
            updated_df = df_new
    else:
        writer = pd.ExcelWriter(output_excel, engine="openpyxl")
        updated_df = df_new

    updated_df.to_excel(writer, sheet_name=freq_name, index=False)
    writer.close()
    print(f"[✓] Sheet '{freq_name}' updated in {output_excel}")

def process_all_tdms(tdms_folder, output_excel):
    for fname in os.listdir(tdms_folder):
        if fname.endswith(".tdms"):
            process_tdms_file(os.path.join(tdms_folder, fname), output_excel)
